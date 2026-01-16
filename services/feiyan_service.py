# -*- coding: utf-8 -*-
"""飞雁数据导入导出与执行结果服务。"""

from __future__ import annotations
import json
import os
import base64
import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import case, func

from extensions.database import db
from models.feiyan_plan_case_result import FeiyanPlanCaseResult
from models.feiyan_test_plan import FeiyanTestPlan
from repositories.feiyan_plan_case_result_repository import FeiyanPlanCaseResultRepository
from repositories.feiyan_test_plan_repository import FeiyanTestPlanRepository
from utils.exceptions import BizError


HEADER_FIELD_MAP = {
    # 飞雁模板表头 -> 系统字段名映射（含历史英文字段兼容）
    "部门ID": "dept_id_ext",
    "部门名称": "dept_name",
    "项目ID": "project_id_ext",
    "项目名称": "project_name",
    "设备ID": "device_id_ext",
    "设备名称": "device_name",
    "测试计划ID": "plan_id_ext",
    "测试计划名称": "plan_name",
    "用例总数": "total",
    "通过": "passed",
    "失败": "failed",
    "阻塞": "blocked",
    "未执行": "not_run",
    "计划测试人员": "tester_names",
    "开始时间": "plan_start_time",
    "结束时间": "plan_end_time",
    "用例分组路径": "group_path",
    "用例ID": "case_id_ext",
    "用例标题": "case_title",
    "优先级": "priority",
    "测试目标": "case_target",
    "用例前置条件": "preconditions",
    "用例执行步骤": "steps_json",
    "用例预期结果": "expected_result",
    "用例关键字": "keywords_json",
    "用例预估执行时间": "workload_minutes",
    "执行人员ID": "executed_by_id",
    "执行人员名称": "executed_by_name",
    "执行开始时间": "execution_start_time",
    "执行结束时间": "execution_end_time",
    "运行结果": "run_result",
    "备注": "remark",
    "失败原因": "failure_reason",
    "BUG编号": "bug_ref",
    "运行结果文件": "attachments_json",
    # 兼容开发模板（英文字段）
    "Department.id": "dept_id_ext",
    "Department.name": "dept_name",
    "Project.id": "project_id_ext",
    "Project.name": "project_name",
    "DeviceModel.id": "device_id_ext",
    "DeviceModel.name": "device_name",
    "TestPlan.id": "plan_id_ext",
    "TestPlan.name": "plan_name",
    "PlanStatistics.total_results": "total",
    "PlanStatistics.passed": "passed",
    "PlanStatistics.failed": "failed",
    "PlanStatistics.blocked": "blocked",
    "PlanStatistics.not_run": "not_run",
    "PlanTester.id": "tester_ids",
    "PlanTester.name": "tester_names",
    "PlanExecutionRun.start_time": "plan_start_time",
    "PlanExecutionRun.end_time": "plan_end_time",
    "PlanCase.group_path": "group_path",
    "PlanCase.case_id": "case_id_ext",
    "PlanCase.title": "case_title",
    "PlanCase.priority": "priority",
    "PlanCase.target": "case_target",
    "PlanCase.preconditions": "preconditions",
    "PlanCase.steps": "steps_json",
    "PlanCase.expected_result": "expected_result",
    "PlanCase.keywords": "keywords_json",
    "PlanCase.workload_minutes": "workload_minutes",
    "CaseExecutionResult.id": "executed_by_id",
    "CaseExecutionResult.name": "executed_by_name",
    "CaseExecutionResult.start_time": "execution_start_time",
    "CaseExecutionResult.end_time": "execution_end_time",
    "CaseExecutionResult.run_result": "run_result",
    "CaseExecutionResult.remark": "remark",
    "CaseExecutionResult.failure_reason": "failure_reason",
    "CaseExecutionResult.bug_ref": "bug_ref",
    "CaseExecutionResult.files": "attachments_json",
}

FEIYAN_TEMPLATE_HEADERS = [
    # 飞雁导入/导出模板列顺序（与模板表头保持一致）
    "部门ID",
    "部门名称",
    "项目ID",
    "项目名称",
    "设备ID",
    "设备名称",
    "测试计划ID",
    "测试计划名称",
    "用例总数",
    "通过",
    "失败",
    "阻塞",
    "未执行",
    "计划测试人员",
    "开始时间",
    "结束时间",
    "用例分组路径",
    "用例ID",
    "用例标题",
    "优先级",
    "测试目标",
    "用例前置条件",
    "用例执行步骤",
    "用例预期结果",
    "用例关键字",
    "用例预估执行时间",
    "执行人员ID",
    "执行人员名称",
    "执行开始时间",
    "执行结束时间",
    "运行结果",
    "备注",
    "失败原因",
    "BUG编号",
    "运行结果文件",
]

REQUIRED_FIELDS = {
    # 导入模板必填字段（用于缺失校验）
    "dept_id_ext": "部门ID",
    "dept_name": "部门名称",
    "project_id_ext": "项目ID",
    "project_name": "项目名称",
    "device_id_ext": "设备ID",
    "plan_id_ext": "测试计划ID",
    "plan_name": "测试计划名称",
    "tester_names": "计划测试人员",
    "case_id_ext": "用例ID",
    "case_title": "用例标题",
    "keywords_json": "用例关键字",
}

ID_FIELDS = {
    "dept_id_ext": "部门ID",
    "project_id_ext": "项目ID",
    "device_id_ext": "设备ID",
    "plan_id_ext": "测试计划ID",
    "case_id_ext": "用例ID",
}

PLAN_FIELDS = {
    "plan_id_ext",
    "dept_id_ext",
    "dept_name",
    "project_id_ext",
    "project_name",
    "plan_name",
    "plan_start_time",
    "plan_end_time",
    "total",
    "passed",
    "failed",
    "blocked",
    "not_run",
    "tester_ids",
    "tester_names",
}

CASE_FIELDS = {
    "plan_id_ext",
    "case_id_ext",
    "case_title",
    "priority",
    "group_path",
    "case_target",
    "preconditions",
    "steps_json",
    "expected_result",
    "keywords_json",
    "workload_minutes",
    "device_id_ext",
    "device_name",
}

RESULT_FIELDS = {
    "executed_by_id",
    "executed_by_name",
    "execution_start_time",
    "execution_end_time",
    "run_result",
    "remark",
    "failure_reason",
    "bug_ref",
    "attachments_json",
}

ALLOWED_RESULTS = {"pass", "fail", "block", "pending", "skip"}


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value).strip() or None


def _normalize_id(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value).strip() or None


def _is_valid_numeric_id(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return value >= 0
    if isinstance(value, float):
        return value.is_integer() and value >= 0
    if isinstance(value, str):
        raw = value.strip()
        return raw.isdigit()
    return False


def _normalize_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return int(float(value))
        except ValueError:
            return None
    return None


def _normalize_time(value: Any, workbook) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value, workbook.epoch)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)
    return str(value).strip() or None


def _normalize_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        if (raw.startswith("{") and raw.endswith("}")) or (
            raw.startswith("[") and raw.endswith("]")
        ):
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return raw
        return raw
    return str(value)


def _normalize_attachments(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        if (raw.startswith("{") and raw.endswith("}")) or (
            raw.startswith("[") and raw.endswith("]")
        ):
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return raw
        parts = [part.strip() for part in raw.split(",") if part.strip()]
        if len(parts) > 1:
            return parts
        return raw
    return value


def _feiyan_storage_root() -> str:
    base_dir = current_app.config.get("ATTACHMENT_STORAGE_DIR")
    if not base_dir:
        base_dir = os.path.join(current_app.root_path, "storage")
    return os.path.join(base_dir, "feiyan")


def _build_attachment_url(file_path: str) -> Optional[str]:
    if not file_path:
        return None
    if file_path.startswith(("http://", "https://")):
        return file_path
    normalized = str(file_path).lstrip("/").replace(os.sep, "/")
    base_url = current_app.config.get("ATTACHMENT_BASE_URL")
    if base_url:
        return urljoin(base_url.rstrip("/") + "/", normalized)
    return f"/api/attachments/{normalized}"


def _store_feiyan_file(
    *,
    file_name: str,
    file_bytes: bytes,
    case_id_ext: str,
    mime_type: Optional[str] = None,
) -> Dict[str, Any]:
    if not file_name:
        raise BizError("附件缺少文件名", 400)
    if not file_bytes:
        raise BizError("附件内容为空", 400)

    safe_file_name = os.path.basename(file_name)
    date_dir = datetime.utcnow().strftime("%Y%m%d")
    storage_root = _feiyan_storage_root()
    case_dir = os.path.join(storage_root, date_dir, case_id_ext)
    os.makedirs(case_dir, exist_ok=True)

    stored_file_name = safe_file_name
    target_path = os.path.join(case_dir, stored_file_name)
    if os.path.exists(target_path):
        stem, ext = os.path.splitext(safe_file_name)
        stored_file_name = f"{stem}_{uuid.uuid4().hex[:8]}{ext}"
        target_path = os.path.join(case_dir, stored_file_name)

    with open(target_path, "wb") as handler:
        handler.write(file_bytes)

    base_dir = current_app.config.get("ATTACHMENT_STORAGE_DIR")
    if not base_dir:
        base_dir = os.path.join(current_app.root_path, "storage")
    rel_path = os.path.relpath(target_path, base_dir).replace(os.sep, "/")
    url = _build_attachment_url(rel_path)
    payload = {
        "file_name": safe_file_name,
        "stored_file_name": stored_file_name,
        "file_path": rel_path,
        "url": url,
        "size": len(file_bytes),
    }
    if mime_type:
        payload["mime_type"] = mime_type
    return payload


def _prepare_feiyan_attachments(attachments: Any, case_id_ext: str) -> Optional[List[Dict[str, Any]]]:
    if attachments is None:
        return None
    normalized = _normalize_attachments(attachments)
    if normalized is None:
        return None
    if isinstance(normalized, list):
        items = normalized
    else:
        items = [normalized]

    prepared: List[Dict[str, Any]] = []
    for item in items:
        if isinstance(item, str):
            prepared.append({"url": item})
            continue
        if not isinstance(item, dict):
            continue

        file_name = item.get("file_name") or item.get("name") or item.get("filename")
        file_bytes = item.get("file_bytes")
        if isinstance(file_bytes, str):
            file_bytes = file_bytes.encode("utf-8")

        if file_bytes:
            prepared.append(
                _store_feiyan_file(
                    file_name=file_name,
                    file_bytes=file_bytes,
                    case_id_ext=case_id_ext,
                    mime_type=item.get("mime_type"),
                )
            )
            continue

        file_content = item.get("content") or item.get("file_content")
        if file_content:
            if not file_name:
                raise BizError("附件缺少文件名", 400)
            content = file_content.split(",", 1)[1] if "," in file_content else file_content
            try:
                raw_bytes = base64.b64decode(content)
            except Exception as exc:  # noqa: BLE001
                raise BizError("附件内容解码失败", 400) from exc
            prepared.append(
                _store_feiyan_file(
                    file_name=file_name,
                    file_bytes=raw_bytes,
                    case_id_ext=case_id_ext,
                    mime_type=item.get("mime_type"),
                )
            )
            continue

        file_path = item.get("file_path")
        if file_path:
            payload = dict(item)
            if not payload.get("url"):
                payload["url"] = _build_attachment_url(file_path)
            prepared.append(payload)
            continue

        url = item.get("url") or item.get("file_url")
        if url:
            payload = dict(item)
            payload["url"] = url
            prepared.append(payload)
            continue

        prepared.append(dict(item))

    return prepared


def _format_attachments(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        urls = []
        for item in value:
            if isinstance(item, dict):
                url = item.get("url") or item.get("file_url")
                if not url:
                    file_path = item.get("file_path")
                    if file_path:
                        url = _build_attachment_url(file_path)
                if url:
                    urls.append(str(url))
            elif isinstance(item, str):
                urls.append(item)
        if urls:
            return ",".join(urls)
    try:
        return json.dumps(value, ensure_ascii=False)
    except TypeError:
        return str(value)


def _format_json_cell(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False)
    except TypeError:
        return str(value)


def _template_path() -> str:
    root = current_app.root_path
    return os.path.join(root, "导入导出数据模版2.xlsx")


class FeiyanService:
    @staticmethod
    def list_departments(page: int, page_size: int) -> Tuple[List[dict], int]:
        return FeiyanTestPlanRepository.list_departments(page, page_size)

    @staticmethod
    def list_test_plans(
        *,
        dept_id_ext: Optional[str],
        project_id_ext: Optional[str],
        page: int,
        page_size: int,
    ) -> Tuple[List[FeiyanTestPlan], int]:
        return FeiyanTestPlanRepository.list_plans(
            dept_id_ext=dept_id_ext,
            project_id_ext=project_id_ext,
            page=page,
            page_size=page_size,
        )

    @staticmethod
    def list_cases(
        *,
        plan_id_ext: str,
        keyword: Optional[str],
        group_path: Optional[str],
        priority: Optional[str],
        run_result: Optional[str],
        page: int,
        page_size: int,
    ) -> Tuple[List[FeiyanPlanCaseResult], int]:
        return FeiyanPlanCaseResultRepository.list_by_plan(
            plan_id_ext=plan_id_ext,
            keyword=keyword,
            group_path=group_path,
            priority=priority,
            run_result=run_result,
            page=page,
            page_size=page_size,
        )

    @staticmethod
    def list_case_group_paths(
        *,
        plan_id_ext: str,
        keyword: Optional[str],
        group_path: Optional[str],
        priority: Optional[str],
        run_result: Optional[str],
    ) -> List[Optional[str]]:
        return FeiyanPlanCaseResultRepository.list_group_paths_by_plan(
            plan_id_ext=plan_id_ext,
            keyword=keyword,
            group_path=group_path,
            priority=priority,
            run_result=run_result,
        )

    @staticmethod
    def import_excel(file_bytes: bytes, current_user=None) -> Dict[str, Any]:
        if not file_bytes:
            raise BizError("导入文件不能为空", 400)

        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        header_row = 1
        if sheet.max_row < header_row:
            raise BizError("Excel模板不完整", 400)

        column_fields: Dict[int, str] = {}

        def _load_headers(row_idx: int):
            # 只识别模板中能映射到系统字段的表头
            mapping: Dict[int, str] = {}
            for idx, cell in enumerate(sheet[row_idx], start=1):
                if cell.value is None:
                    continue
                header = str(cell.value).strip()
                field = HEADER_FIELD_MAP.get(header)
                if field:
                    mapping[idx] = field
            return mapping

        # 读取模板表头并建立列索引映射
        column_fields = _load_headers(header_row)
        if not column_fields:
            raise BizError("Excel缺少有效字段列", 400)
        present_fields = set(column_fields.values())
        missing_headers = [
            display for field, display in REQUIRED_FIELDS.items() if field not in present_fields
        ]
        if missing_headers:
            raise BizError(f"Excel缺少必填字段: {', '.join(missing_headers)}", 400)

        plan_cache: Dict[str, FeiyanTestPlan] = {}
        case_cache: Dict[str, FeiyanPlanCaseResult] = {}
        touched_plans: set[str] = set()

        success_count = 0
        failure_count = 0
        errors: List[Dict[str, Any]] = []

        data_start_row = header_row + 1
        for row_idx in range(data_start_row, sheet.max_row + 1):
            raw_row: Dict[str, Any] = {}
            for col_idx, field in column_fields.items():
                raw_row[field] = sheet.cell(row=row_idx, column=col_idx).value

            if all(_is_blank(value) for value in raw_row.values()):
                continue
            try:
                missing_values = [
                    display
                    for field, display in REQUIRED_FIELDS.items()
                    if _is_blank(raw_row.get(field))
                ]
                if missing_values:
                    raise BizError(
                        f"第{row_idx}行必填字段为空: {', '.join(missing_values)}", 400
                    )
                plan_payload = FeiyanService._build_plan_payload(raw_row, workbook)
                case_payload, result_payload = FeiyanService._build_case_payload(raw_row, workbook)
                plan_id_ext = plan_payload.get("plan_id_ext")
                case_id_ext = case_payload.get("case_id_ext")
                if not plan_id_ext:
                    raise BizError(f"第{row_idx}行缺少计划ID", 400)
                if not case_id_ext:
                    raise BizError(f"第{row_idx}行缺少用例ID", 400)

                plan = plan_cache.get(plan_id_ext)
                if plan is None:
                    plan = FeiyanTestPlanRepository.get_by_plan_id(plan_id_ext)
                    if plan is None:
                        plan = FeiyanTestPlan(
                            plan_id_ext=plan_id_ext,
                            created_by_user_id=(current_user.id if current_user else None),
                        )
                        FeiyanTestPlanRepository.add(plan)
                    plan_cache[plan_id_ext] = plan

                FeiyanService._apply_updates(plan, plan_payload, allow_blank=False)
                touched_plans.add(plan_id_ext)

                case = case_cache.get(case_id_ext)
                if case is None:
                    case = FeiyanPlanCaseResultRepository.get_by_case_id(case_id_ext)
                    if case is None:
                        case = FeiyanPlanCaseResult(
                            plan_id_ext=plan_id_ext,
                            case_id_ext=case_id_ext,
                        )
                        FeiyanPlanCaseResultRepository.add(case)
                    case_cache[case_id_ext] = case

                case.plan_id_ext = plan_id_ext
                FeiyanService._apply_updates(case, case_payload, allow_blank=False)

                if not _is_blank(result_payload.get("run_result")):
                    FeiyanService._apply_updates(case, result_payload, allow_blank=True)

                success_count += 1
            except BizError as exc:
                failure_count += 1
                errors.append({"row": row_idx, "message": exc.message})

        try:
            if touched_plans:
                for plan_id_ext in touched_plans:
                    FeiyanService.refresh_plan_statistics(plan_id_ext)

            FeiyanTestPlanRepository.commit()
            FeiyanPlanCaseResultRepository.commit()
        except Exception:
            FeiyanTestPlanRepository.rollback()
            FeiyanPlanCaseResultRepository.rollback()
            raise

        return {
            "success_count": success_count,
            "failure_count": failure_count,
            "errors": errors,
        }

    @staticmethod
    def record_result(plan_id_ext: str, payload: Dict[str, Any], current_user=None) -> FeiyanPlanCaseResult:
        plan_id_ext = _normalize_id(plan_id_ext)
        case_id_ext = _normalize_id(payload.get("case_id_ext") or payload.get("case_id"))
        if not plan_id_ext:
            raise BizError("plan_id 不能为空", 400)
        if not case_id_ext:
            raise BizError("case_id 不能为空", 400)

        case_result = FeiyanPlanCaseResultRepository.get_by_plan_and_case(plan_id_ext, case_id_ext)
        if not case_result:
            raise BizError("用例不存在", 404)

        run_result = _normalize_text(payload.get("run_result"))
        if run_result:
            run_result = run_result.lower()
        if _is_blank(run_result):
            return case_result
        if run_result not in ALLOWED_RESULTS:
            raise BizError(f"run_result 必须是 {sorted(ALLOWED_RESULTS)} 之一", 400)

        result_payload = {
            "run_result": run_result,
            "remark": _normalize_text(payload.get("remark")),
            "failure_reason": _normalize_text(payload.get("failure_reason")),
            "bug_ref": _normalize_text(payload.get("bug_ref")),
            "execution_start_time": _normalize_text(payload.get("execution_start_time")),
            "execution_end_time": _normalize_text(payload.get("execution_end_time")),
        }

        if "attachments" in payload or "attachments_json" in payload:
            attachments = _prepare_feiyan_attachments(
                payload.get("attachments") or payload.get("attachments_json"),
                case_id_ext=case_id_ext,
            )
            result_payload["attachments_json"] = attachments

        executed_by_name = _normalize_text(payload.get("executed_by_name"))
        executed_by_id = _normalize_text(payload.get("executed_by_id"))
        if _is_blank(executed_by_name) and current_user is not None:
            executed_by_name = getattr(current_user, "username", None)
        if _is_blank(executed_by_id) and current_user is not None:
            executed_by_id = str(getattr(current_user, "id", ""))

        result_payload["executed_by_name"] = executed_by_name
        result_payload["executed_by_id"] = executed_by_id

        FeiyanService._apply_updates(case_result, result_payload, allow_blank=True)
        try:
            FeiyanService.refresh_plan_statistics(plan_id_ext)
            FeiyanPlanCaseResultRepository.commit()
            FeiyanTestPlanRepository.commit()
        except Exception:
            FeiyanPlanCaseResultRepository.rollback()
            FeiyanTestPlanRepository.rollback()
            raise
        return case_result

    @staticmethod
    def export_plan(plan_id_ext: str) -> bytes:
        plan_id_ext = _normalize_id(plan_id_ext)
        if not plan_id_ext:
            raise BizError("plan_id 不能为空", 400)

        plan = FeiyanTestPlanRepository.get_by_plan_id(plan_id_ext)
        if not plan:
            raise BizError("测试计划不存在", 404)

        FeiyanService.refresh_plan_statistics(plan_id_ext)
        plan = FeiyanTestPlanRepository.get_by_plan_id(plan_id_ext)
        if not plan:
            raise BizError("测试计划不存在", 404)

        case_results, _ = FeiyanPlanCaseResultRepository.list_by_plan(
            plan_id_ext=plan_id_ext,
            page=1,
            page_size=10_000_000,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数据导入导出模板"

        header_row = 1
        for idx, header in enumerate(FEIYAN_TEMPLATE_HEADERS, start=1):
            sheet.cell(row=header_row, column=idx, value=header)

        data_start_row = header_row + 1

        field_by_col: Dict[int, str] = {}
        for idx, header in enumerate(FEIYAN_TEMPLATE_HEADERS, start=1):
            field = HEADER_FIELD_MAP.get(header)
            if field:
                field_by_col[idx] = field

        def _resolve_value(field: str, item: FeiyanPlanCaseResult) -> Any:
            if field in PLAN_FIELDS:
                return getattr(plan, field, None)
            if field == "attachments_json":
                return _format_attachments(item.attachments_json)
            if field in ("steps_json", "keywords_json"):
                return _format_json_cell(getattr(item, field, None))
            if field == "case_id_ext":
                return item.case_id_ext
            if field == "case_title":
                return item.case_title
            if hasattr(item, field):
                return getattr(item, field)
            return None

        for row_offset, item in enumerate(case_results):
            row_idx = data_start_row + row_offset
            for col_idx, field in field_by_col.items():
                value = _resolve_value(field, item)
                sheet.cell(row=row_idx, column=col_idx, value=value)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def refresh_plan_statistics(plan_id_ext: str):
        plan = FeiyanTestPlanRepository.get_by_plan_id(plan_id_ext)
        if not plan:
            return

        result_stmt = db.session.query(
            func.count(FeiyanPlanCaseResult.id),
            func.sum(case((FeiyanPlanCaseResult.run_result == "pass", 1), else_=0)),
            func.sum(case((FeiyanPlanCaseResult.run_result == "fail", 1), else_=0)),
            func.sum(case((FeiyanPlanCaseResult.run_result == "block", 1), else_=0)),
            func.sum(
                case(
                    (
                        (FeiyanPlanCaseResult.run_result.is_(None))
                        | (FeiyanPlanCaseResult.run_result == "")
                        | (FeiyanPlanCaseResult.run_result == "pending")
                        | (FeiyanPlanCaseResult.run_result == "skip"),
                        1,
                    ),
                    else_=0,
                )
            ),
        ).filter(FeiyanPlanCaseResult.plan_id_ext == plan_id_ext)

        total, passed, failed, blocked, not_run = result_stmt.one()
        plan.total = int(total or 0)
        plan.passed = int(passed or 0)
        plan.failed = int(failed or 0)
        plan.blocked = int(blocked or 0)
        plan.not_run = int(not_run or 0)

    @staticmethod
    def _build_plan_payload(raw_row: Dict[str, Any], workbook) -> Dict[str, Any]:
        payload: Dict[str, Any] = {}
        for field in PLAN_FIELDS:
            if field not in raw_row:
                continue
            value = raw_row.get(field)
            if field in {"plan_id_ext", "dept_id_ext", "project_id_ext"}:
                payload[field] = _normalize_id(value)
            elif field in {"total", "passed", "failed", "blocked", "not_run"}:
                payload[field] = _normalize_int(value)
            elif field in {"plan_start_time", "plan_end_time"}:
                payload[field] = _normalize_time(value, workbook)
            elif field == "tester_names":
                payload[field] = _normalize_text(value)
            elif field == "tester_ids":
                payload[field] = _normalize_json_value(value)
            else:
                payload[field] = _normalize_text(value)
        return payload

    @staticmethod
    def _build_case_payload(
        raw_row: Dict[str, Any],
        workbook,
    ) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        case_payload: Dict[str, Any] = {}
        result_payload: Dict[str, Any] = {}

        for field in CASE_FIELDS:
            if field not in raw_row:
                continue
            value = raw_row.get(field)
            if field in {"plan_id_ext", "case_id_ext", "device_id_ext"}:
                case_payload[field] = _normalize_id(value)
            elif field == "workload_minutes":
                case_payload[field] = _normalize_int(value)
            elif field in {"steps_json", "keywords_json"}:
                case_payload[field] = _normalize_json_value(value)
            else:
                case_payload[field] = _normalize_text(value)

        for field in RESULT_FIELDS:
            if field not in raw_row:
                continue
            value = raw_row.get(field)
            if field == "run_result":
                normalized = _normalize_text(value)
                if normalized:
                    normalized = normalized.lower()
                result_payload[field] = normalized
            elif field in {"execution_start_time", "execution_end_time"}:
                result_payload[field] = _normalize_time(value, workbook)
            elif field == "attachments_json":
                result_payload[field] = _normalize_attachments(value)
            else:
                result_payload[field] = _normalize_text(value)

        if "run_result" in result_payload and not _is_blank(result_payload["run_result"]):
            if result_payload["run_result"] not in ALLOWED_RESULTS:
                raise BizError(f"run_result 必须是 {sorted(ALLOWED_RESULTS)} 之一", 400)

        return case_payload, result_payload

    @staticmethod
    def _apply_updates(target, payload: Dict[str, Any], *, allow_blank: bool):
        for key, value in payload.items():
            if not allow_blank and _is_blank(value):
                continue
            setattr(target, key, value)
